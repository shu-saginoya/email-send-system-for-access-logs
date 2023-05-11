from http import server
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import tkinter.messagebox as mb
import tkinter.filedialog as fd
import smtplib
import openpyxl
import sys
import os
import re

#ダイアログなどに表示されるタイトル
PROGRAM_NAME = "一斉メール送信システム"

#対応エクセルファイル名
EXCEL_NAME = 'send_mail_list.xlsx'

#対応エクセルファイルのシート名
SHEET_SETTINGS = 'Send settings'
SHEET_CONTENTS = 'Send contents'

#対応エクセルファイルのセルの位置を指定
#シート1
FROM_ADDRESS_CELL = [2, 3]
SMTP_SERVER_CELL = [3, 3]
SMTP_PORT_CELL = [4, 3]
SMTP_USER_CELL = [5, 3]
SMTP_PASSWORD_CELL = [6, 3]
READ_START_ROW_NO = 9
TO_NAME_COL_NO = 2
TO_ADDRESS_COL_NO = 3
ATTACHMENT_INDIVIDUAL_COL_NO = 4
#シート2
MAIL_SUBJECT_CELL = [2,3]
MAIL_BODY_CELL = [3,3]
ATTACHMENT_COMMON_CELL = [4,3]

#正規表現
PATTERN_MAIL = "^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

#設定ファイル・添付ファイルのパスを返す関数
def file_path(fail_name):
    if directory_path:
        return directory_path + '/' + fail_name

#実行の確認
execution_confirmation = mb.askyesno(
    PROGRAM_NAME,
    f'''メールを一斉送信します。
設定用エクセルファイル（{EXCEL_NAME}）とすべての添付ファイルを同じディレクトリに準備してください。
次の操作でそのディレクトリを選択します。
実行してよろしいですか？'''
)
if execution_confirmation == False:
    mb.showinfo(PROGRAM_NAME, '処理は中止されました。')
    sys.exit()
    
#対応エクセルから値の取得
try:
    #設定ファイル・添付ファイルの格納されたディレクトリをユーザーに入力させる
    directory_path = fd.askdirectory(
        title='設定ファイル・添付ファイルのフォルダーを選択'
    )
    #設定ファイルから値の取得
    #設定ファイルの有無をチェック
    if os.path.isfile(file_path(EXCEL_NAME)):
        book = openpyxl.load_workbook(file_path(EXCEL_NAME))
    else:
        raise FileNotFoundError(EXCEL_NAME + 'が見つかりません。')
    sheet_send_settings = book[ SHEET_SETTINGS ]
    sheet_send_contents = book[ SHEET_CONTENTS ]
    from_address = sheet_send_settings.cell(*FROM_ADDRESS_CELL).value
    if not from_address:
        raise ValueError('送信元メールアドレスが未入力です。')
    to_address_list = []
    attachment_common = sheet_send_contents.cell(*ATTACHMENT_COMMON_CELL).value
    if os.path.isfile(file_path(attachment_common)) == False:
        raise FileNotFoundError(attachment_common + 'が見つかりません。')
    attachment_individual_list = []

    smtp_host = sheet_send_settings.cell(*SMTP_SERVER_CELL).value
    smtp_port = sheet_send_settings.cell(*SMTP_PORT_CELL).value
    smtp_user = sheet_send_settings.cell(*SMTP_USER_CELL).value
    smtp_pass = sheet_send_settings.cell(*SMTP_PASSWORD_CELL).value
    if not smtp_host or not smtp_port or not smtp_user or not smtp_pass:
        raise ValueError('SMTPサーバーの設定に未入力があります。')

    contents_subject = sheet_send_contents.cell(*MAIL_SUBJECT_CELL).value
    contents_body_before_correction = sheet_send_contents.cell(*MAIL_BODY_CELL).value
    if not contents_subject or not contents_body_before_correction:
        raise ValueError('メールの件名・本文に未入力があります。')
    contents_body = contents_body_before_correction.replace('\n', '<br>')

    for row_no in range(READ_START_ROW_NO, sheet_send_settings.max_row+1):
        to_address = sheet_send_settings.cell(row_no, TO_ADDRESS_COL_NO).value
        to_attachment_individual = sheet_send_settings.cell(row_no, ATTACHMENT_INDIVIDUAL_COL_NO).value
        if to_address is None:
            break
        elif re.match(PATTERN_MAIL, to_address):
            to_address_list.append(to_address)
        else:
            raise ValueError(to_address + 'のメールアドレスが正しくありません。')
        if os.path.isfile(file_path(to_attachment_individual)):
            attachment_individual_list.append(to_attachment_individual)
        else:
            raise FileNotFoundError(to_attachment_individual + 'が見つかりません。')

    #SMTPサーバーの設定
    server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.set_debuglevel(0)
    server.login(smtp_user, smtp_pass)

    #メールの送信
    for i in range(len(to_address_list)):
        message = MIMEMultipart()
        message['Subject'] = contents_subject
        message['From'] = from_address
        message['To'] = to_address_list[i]
        message.attach(MIMEText(contents_body, 'html'))

        attachment_list = [attachment_common, attachment_individual_list[i]]
        for file in attachment_list:
            with open(file_path(file), "rb") as f:
                attachment = MIMEApplication(f.read())

            attachment.add_header(
                "Content-Disposition",
                "attachment",
                filename=file
            )
            message.attach(attachment)

        server.send_message(message)

    #SMTPサーバーの終了
    server.quit()

except FileNotFoundError as err:
    mb.showerror(PROGRAM_NAME, f'''必要なファイルが見つかりませんでした。
設定エクセルファイルおよび添付ファイルはひとつのディレクトリにまとめて設置してください。
{err}''')
except ValueError as err:
    mb.showerror(PROGRAM_NAME, f'''不正な値が代入されました。
設定ファイルに誤りがないか確認してください。
{err}''')
except smtplib.SMTPAuthenticationError as err:
    mb.showerror(PROGRAM_NAME, f'''SMTPサーバーへの接続に失敗しました。
{err}''')
except Exception as err:
    mb.showerror(PROGRAM_NAME, f'''エラーが発生しました。
{err}''')
else:
    mb.showinfo(PROGRAM_NAME, 'メールの送信が完了しました。')